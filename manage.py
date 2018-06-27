#!/usr/bin/env python
import os
from openpyxl import load_workbook

from app import create_app, db
from app.models import User, Follow, Role, Permission, Post, Comment, Word, WordInterpretation
from flask_script import Manager, Shell
from flask_migrate import Migrate, MigrateCommand

app = create_app(os.getenv('FLASK_CONFIG') or 'default')
manager = Manager(app)
migrate = Migrate(app, db, compare_type=True) # 检查类型, 字段长度的变化


def make_shell_context():
    return dict(app=app, db=db, User=User, Follow=Follow, Role=Role,
                Permission=Permission, Post=Post, Comment=Comment)
manager.add_command("shell", Shell(make_context=make_shell_context))
manager.add_command('db', MigrateCommand)


def command_list_routes(name):
    from colorama import init, Fore
    from tabulate import tabulate
    init()
    table = []
    for rule in app.url_map.iter_rules():
        table.append([
            Fore.BLUE + rule.endpoint,
            Fore.GREEN + ','.join(rule.methods),
            Fore.YELLOW + rule.rule])

    print(tabulate(sorted(table),
                   headers=(
                       Fore.BLUE + 'End Point(method name)',
                       Fore.GREEN + 'Allowed Methods',
                       Fore.YELLOW + 'Routes'
                   ), showindex="always", tablefmt="grid"))


@manager.command
def routes():
    command_list_routes('API Routes')

@manager.command
def import_words():
    wb = load_workbook(filename='init_data/words.xlsx')
    ws = wb.active
    rows = ws.max_row
    last_word = None
    for row in range(2, rows):
        try:
            word = ws['A{}'.format(row)].value
            type = ws['B{}'.format(row)].value
            interpretation = ws['C{}'.format(row)].value
            """
            word, type, interpretation, word not duplicated: pass : 标准
            , type, interpretation, last_word: pass
            , type, interpretation, last_word=None: pass
            , , , ignore
            """
            if (word and word.strip()) and (type and type.strip()) and ( interpretation and interpretation.strip()) \
                and Word.query.filter(Word.word == word).count() < 1:
                word = word.strip()
                type = type.strip()
                interpretation = interpretation.strip()
                word_obj = Word()
                word_obj.word = word
                db.session.add(word_obj)
                db.session.flush()
                last_word = word_obj

                word_interpretation = WordInterpretation()
                word_interpretation.word_id = word_obj.id
                word_interpretation.type = type
                word_interpretation.interpretation = interpretation
                db.session.add(word_interpretation)

            elif (type and type.strip()) and (interpretation and interpretation.strip()) \
                and last_word:
                type = type.strip()
                interpretation = interpretation.strip()
                word_interpretation = WordInterpretation()
                word_interpretation.word_id = last_word.id
                word_interpretation.type = type
                word_interpretation.interpretation = interpretation
                db.session.add(word_interpretation)
            else:
                pass
        except Exception as e:
            print('{}行数据有问题!!'.format(row))
            raise(e)
    db.session.commit()


@manager.command
def deploy():
    """Run deployment tasks."""
    from flask_migrate import upgrade
    from app.models import Role, User

    # migrate database to latest revision
    upgrade()

    # create user roles
    Role.insert_roles()

    # create self-follows for all users
    User.add_self_follows()


if __name__ == '__main__':
    manager.run()
